from openpyxl import Workbook, load_workbook
import os
import time
import openpyxl
from openpyxl.styles import Alignment, Border, Side, PatternFill, Font, Color, colors
from difflib import SequenceMatcher

def check_inputs(file):
    if "DUMMY" in file:
        return 0
    else:
        if "MT" in file:
            if  "2MT" in file or "MT2" in file:
                if not "ISG" in file:
                    return 1
                else:
                    if "no" in file:
                        return 2
                    else:
                        return 3
            else:
                if not "ISG" in file:
                    return 4
                else:
                    if "no" in file:
                        return 5
                    else:
                        return 6

        else:
            if "2AT" in file or "AT2" in file:
                if not "ISG" in file:
                    return 7
                else:
                    if "no" in file:
                        return 8
                    else:
                        return 9
            else:
                if not "ISG" in file:
                    return 10
                else:
                    if "no" in file:
                        return 11
                    else:
                        return 12

def check_variant(file):

    if "DUMMY" in file:
        return 0
    else:
        if "MT" in file:
            if  "2WD" in file or "WD2" in file:
                if not "ISG" in file:
                    return 1
                else:
                    if "no" in file:
                        return 2
                    else:
                        return 3
            else:
                if not "ISG" in file:
                    return  4
                else:
                    if "no" in file:
                        return 5
                    else:
                        return 6

        else:
            if "2WD" in file or "WD2" in file:
                if not "ISG" in file:
                    return 7
                else:
                    if "no" in file:
                        return 8
                    else:
                        return 9
            else:
                if not "ISG" in file:
                    return 10
                else:
                    if "no" in file:
                        return 11
                    else:
                        return 12

def set_border_style(style):
    return Border(left=Side(style=style), right=Side(style=style), top=Side(style=style), bottom=Side(style=style))

def set_aligment(horizontal, vertical):
    return openpyxl.styles.Alignment(horizontal=horizontal,vertical=horizontal)


date = time.strftime("%d.%m.%Y")

calid_list = []
cvn_list = []
ordered_ulp_list = []
check_list = []
calibration_tab_names = []
input_ulp_files = []

vehicle_reference = ""
customer_part_number = ""

vehicle_variants_xl ="SYMC Vehicle Variants.xlsx"
m_files_xl = "Example_Worksheet in M-files tests.xlsx"
tracking_xl = "M_files_tracking.xlsm"

merged_row = 9
missing_value = False

tracking_tab_name = "M file history"
variant_tab_name = "X117"

input_folder = "/Input"
output_folder = "/Output"
empty = "Empty"

who = "Selim"
m_file_aim = "Deneme"

vehicle_variants_table = load_workbook(vehicle_variants_xl, data_only=True)
vv_sheet = vehicle_variants_table[variant_tab_name]

m_files_table = load_workbook(m_files_xl, data_only=True)

tracking_table = load_workbook(tracking_xl, data_only=True, read_only=False, keep_vba=True)
tracking_sheet = tracking_table[tracking_tab_name]

if not os.path.isfile(vehicle_variants_xl):
    print vehicle_variants_xl + " not found."
elif not os.path.isfile(m_files_xl):
    print m_files_xl + " not found."
elif not os.path.isfile(tracking_xl):
    print tracking_xl + " not found."
else:
    folders = filter(lambda x: os.path.isdir(x), os.listdir('.'))
    if not len (folders) is 0:

        print "Please enter your name:"
        who = raw_input()
        print "Please enter mfiles aim:"
        m_file_aim = raw_input()

        tab_found = False
        while True:
            print "Please enter vehicle variant:" 
            tab_input = raw_input()
            for sheets in vehicle_variants_table.worksheets:
                if tab_input == sheets.title:
                    variant_tab_name = tab_input
                    tab_found = True
            if tab_found:
                break
            else:
                print tab_input + " not found. Please check tab name."
        print"a"

        for folder in folders:
            c_p_n = folder.split("_")[-1]
            customer_part_number = c_p_n[0:3] + " " + c_p_n[3:6] + " " + c_p_n[6:8] + " " + c_p_n[8:10]

            for name in os.listdir(folder + input_folder):
                if name.endswith('.ulp'):
                    input_ulp_files.append(name)

            for name in os.listdir(folder + output_folder):
                if name.endswith('.ulp'):
                    output_ulp_files = name

            sw_name = folder.split("_")[0]

            is_found = False

            for i in range(1, vv_sheet.max_row):
                if vv_sheet.cell(row=i, column=3).value == customer_part_number:
                    vehicle_reference = vv_sheet.cell(row=i, column=1).value
                    is_found = True
                    break

            if is_found:
                for c in range(i,i+merged_row):
                    if vv_sheet.cell(row=c, column=6).value == empty:
                        calid_list.append(empty)
                        cvn_list.append(empty)
                        calibration_tab_names.append(empty)
                    else:
                        calibration_tab_name = vv_sheet.cell(row=c, column=6).value
                        calibration_tab_names.append(calibration_tab_name)
                        m_files_sheet = m_files_table[calibration_tab_name]
                        cid = m_files_sheet.cell(row=17, column=3).value
                        cvn = m_files_sheet.cell(row=18, column=3).value

                        if cid == "" or cid == " ":
                            print "Calibration id not found on " + calibration_tab_name + " named tab."
                            missing_value = True
                            break
                        elif cvn == "" or cvn == " ":
                            print "Cvn number not found on " + calibration_tab_name + " named tab."
                            missing_value = True
                            break
                        else:
                            calid_list.append(cid)
                            cvn_list.append(cvn)

                for tab in calibration_tab_names:
                    if tab == empty:
                        ordered_ulp_list.append(empty)
                    else:
                        for file in input_ulp_files:
                            if check_variant(tab) == check_inputs(file):
                                ordered_ulp_list.append(file)

                for row in range (0,merged_row):
                    tracking_sheet.append((m_file_aim, vehicle_reference, customer_part_number,sw_name, date, who, output_ulp_files, ordered_ulp_list[row], row, cvn_list[row], '', calid_list[row] ))
                    ts_max = tracking_sheet.max_row
                    for let in range(1, 14):
                        tracking_sheet.cell(row=ts_max, column= let).alignment = set_aligment('center','center')
                        if let<7:
                            tracking_sheet.cell(row=ts_max, column=let).font = Font(bold=True)
                        if tracking_sheet.cell(row=ts_max, column=8).value == empty:
                            for col in (8,10,12):
                                tracking_sheet.cell(row=ts_max, column=col).font = Font(bold=True, color=colors.RED)

                for c in "ABCDEFG":
                    tracking_sheet.merge_cells(c + str(ts_max-8) + ':' + c + str(ts_max))

                calid_list = []
                cvn_list = []
                input_ulp_files = []
                calibration_tab_names= []
                ordered_ulp_list = []

            else:
                print customer_part_number + " customer number not found in " + vv_sheet
                missing_value = True
                break



        for row in range(7,ts_max+1):
            for col in range(1,14):
                tracking_sheet.cell(row=row, column=col).border = set_border_style('thin')

        if not missing_value:
            tracking_table.save(tracking_xl)
            print "Process completed."

    else:
        print "There is a no folder."